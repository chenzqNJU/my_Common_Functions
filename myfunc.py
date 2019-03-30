def search(key, dirFlag=0, file_dir=r'C:\Users\chenzq\Documents\PycharmProjects\code'):
    import os
    temp = []
    for root, dirs, files in os.walk(file_dir):
        for file in files:
            if os.path.splitext(file)[1] in ['.py', '.txt']:
                temp.append(os.path.join(root, file))
    file_dir = r'C:\Users\chenzq\Documents\PycharmProjects'
    files = os.listdir(file_dir)
    for fi in files:
        fi_d = os.path.join(file_dir, fi)
        if os.path.splitext(fi_d)[1] in ['.py', '.txt']:
            temp.append(fi_d)
    for dir in temp:
        s = 0
        # dir=file[8]
        try:
            with open(dir, encoding='utf-8') as f:
                for i in f:
                    if key in i:
                        print(i),
                        s += 1
                        if dirFlag == 1:
                            print(dir)
                # if s == 0:
                # print("don't match it！")
        except:
            try:
                with open(dir, encoding='gbk') as f:
                    for i in f:
                        if key in i:
                            print(i),
                            s += 1
                            if dirFlag == 1:
                                print(dir)
                    # if s == 0:
                    # print("don't match it！")
            except:
                print(dir)

    # import os
    # def gci(filepath):
    #     filepath=file_dir
    #     # 遍历filepath下所有文件，包括子目录
    #     files = os.listdir(filepath)
    #     fi= files[2]
    #     for fi in files:
    #         fi_d = os.path.join(filepath, fi)
    #         if os.path.isdir(fi_d):
    #             gci(fi_d)
    #         else:
    #             print(os.path.join(filepath, fi_d))


def delete():
    z = list(globals().keys())
    print(z)
    for key in z:
        if not key.startswith("__"):
            print(key)
            globals().pop(key)


def __Delete(t):
    z = list(globals().keys())
    print(z)
    z1 = len(t)
    for key in z:
        if key.startswith(t) and ((len(key) == z1) or not key[z1].isalpha()):
            print(key)
            globals().pop(key)


def add(x, y):
    z = x * y
    return z


import sys, time


class ShowProcess():
    import sys, time

    """
    显示处理进度的类
    调用该类相关函数即可实现处理进度的显示
    """
    i = 1  # 当前的处理进度
    max_steps = 0  # 总共需要处理的次数
    max_arrow = 50  # 进度条的长度

    # 初始化函数，需要知道总共的处理次数
    def __init__(self, max_steps):
        self.max_steps = max_steps
        self.i = 1

    # 显示函数，根据当前的处理进度i显示进度
    # 效果为[>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>]100.00%
    def show_process(self, i=None):
        import numpy as np
        if i is not None:
            self.i = i
        num_arrow = int(self.i * self.max_arrow / self.max_steps)  # 计算显示多少个'>'
        num_line = self.max_arrow - num_arrow  # 计算显示多少个'-'
        percent = self.i * 100.0 / self.max_steps  # 计算完成进度，格式为xx.xx%
        process_bar = '[' + '>' * num_arrow + '-' * num_line + ']' \
                      + '%.2f' % percent + '%' + '\r'  # 带输出的字符串，'\r'表示不换行回到最左边
        if np.floor(percent) != np.floor(percent - 100.0 / self.max_steps):
            sys.stdout.write(process_bar)  # 这两句打印字符到终端
            sys.stdout.flush()
        self.i += 1

    def close(self, words='done'):
        print('')
        print(words)
        self.i = 1

    '''
    if __name__=='__main__':
        max_steps = 100
    
        process_bar = ShowProcess(max_steps)
    
        for i in range(max_steps + 1):
            process_bar.show_process()
           # time.sleep(0.05)
        process_bar.close()   '''


def flag_jishu(df, key='date'):
    import numpy as np
    t = df[key].values
    t1 = t[1:] > t[:-1]
    t1 = np.concatenate(([True], t1))
    t2 = np.cumsum(t1)
    df['flag'] = t2


def zhangdie(Price):
    import numpy as np
    temp = Price[:1].copy();
    temp[:] = np.nan
    temp1 = round((Price + 0.000001) * 1.1, 2)
    temp1 = temp.append(temp1[:-1])
    temp1.index = Price.index
    zhangting = temp1 == Price
    temp2 = round((Price + 0.000001) * 0.9, 2)
    temp2 = temp.append(temp2[:-1])
    temp2.index = Price.index
    dieting = (temp2 == Price)
    return zhangting.astype(int) - dieting.astype(int)


def kxiantu(df, code=''):
    from matplotlib import dates as mdates
    import datetime as dt
    from matplotlib.finance import candlestick_ohlc
    from matplotlib import ticker as mticker
    import pandas as pd
    import matplotlib.pyplot as plt
    plt.close('all')
    t = df.columns
    date = t[t.str.contains('ate')][0]
    open = t[t.str.contains('pen')][0]
    high = t[t.str.contains('igh')][0]
    low = t[t.str.contains('ow')][0]
    close = t[t.str.contains('lose')][0]
    t = df[date].reset_index(drop=True)  # 标签
    df[date] = pd.to_datetime(df[date])
    df.loc[:, 'DateTime'] = mdates.date2num(df.loc[:, date].astype(dt.date))
    df.drop(date, axis=1, inplace=True)
    # 调整顺序
    df = df.reindex(columns=['DateTime', open, high, low, close])
    SP = len(df)
    fig = plt.figure(facecolor='#07000d', figsize=(15, 10))
    ax1 = plt.subplot2grid((6, 4), (1, 0), rowspan=4, colspan=4, axisbg='#07000d')
    # 蜡烛图
    t1 = df.iloc[0, 0]
    df.DateTime = range(len(df)) + t1
    candlestick_ohlc(ax1, df.values[-SP:], width=0.8, colorup='#ff1717', colordown='#53c156')

    ax1.grid(True, color='blue')
    ax1.xaxis.set_major_locator(mticker.MaxNLocator(10))
    ax1.xaxis.set_major_formatter(mdates.DateFormatter('%Y-%m-%d'))
    ax1.yaxis.label.set_color("blue")

    #######################避免k线出现断裂（周末）采用label标签的方法
    label = []
    for x in ax1.get_xticks():
        if x >= t1 and x <= t1 + len(df) - 1:
            label += [t[x - t1]]
        elif x <= t1 + len(df) - 1:
            label += [mdates.num2date(x).strftime("%Y-%m-%d")]
        else:
            label += ['']
    ax1.set_xticklabels(label)
    # 上下左右的边框线为蓝色
    ax1.spines['bottom'].set_color("#5998ff")
    ax1.spines['top'].set_color("#5998ff")
    ax1.spines['left'].set_color("#5998ff")
    ax1.spines['right'].set_color("#5998ff")
    ax1.tick_params(axis='y', colors='blue')
    plt.gca().yaxis.set_major_locator(mticker.MaxNLocator(prune='upper'))
    ax1.tick_params(axis='x', colors='blue')
    plt.ylabel('Stock price and Volume')
    plt.title(code, fontsize='large', color='blue')
    return fig
    plt.show()

def kdj(df):
    # df=data
    import pandas as pd
    t = df.columns
    date = t[t.str.contains('ate')][0]
    open = t[t.str.contains('pen')][0]
    high = t[t.str.contains('igh')][0]
    low = t[t.str.contains('ow')][0]
    close = t[t.str.contains('lose')][0]
    ticker = ['code', 'ticker'][t.str.contains('ticker').any()]
    t = df[date].reset_index(drop=True)
    df[date] = pd.to_datetime(df[date])

    Price = pd.pivot_table(df, index=date, columns=ticker, values=close)

    for i in ['openPrice', 'lowestPrice', 'highestPrice']:
        locals()[i] = pd.read_csv(i + '.csv').iloc[-100:, :10]
        locals()[i].tradeDate = pd.to_datetime(locals()[i].tradeDate, format='%Y-%m-%d')
        locals()[i].set_index('tradeDate', inplace=True)
        locals()[i] = locals()[i].loc[Price.index]

    RSV = (Price - lowestPrice) / (highestPrice - lowestPrice) * 100
    RSV.replace(np.inf, np.nan, inplace=True)
    # 涨跌停
    temp = Price[:1].copy();
    temp[:] = np.nan
    temp1 = round((Price + 0.000001) * 1.1, 2)
    temp1 = temp.append(temp1[:-1])
    temp1.index = Price.index
    zhangting = temp1 == Price
    temp2 = round((Price + 0.000001) * 0.9, 2)
    temp2 = temp.append(temp2[:-1])
    temp2.index = Price.index
    dieting = (temp2 == Price)

    RSV[dieting] = 0
    RSV[zhangting] = 100
    ####计算k值
    for n, i in enumerate(RSV.index):
        # n=1;i=Price.index[n]
        if n == 0:
            K = np.ones((1, RSV.shape[1])) * 50;
            tempK = K[0].copy()
            D = K.copy()
            tempD = tempK.copy()
            RSV_ = RSV.values
        else:
            x = np.where(isOpen.loc[i] == 1)
            tempK[x] = RSV_[n, x] * 1 / 3 + tempK[x] * 2 / 3
            # temp[np.isnan(temp)^np.isnan(Price_[n,])]=50
            K = np.row_stack((K, tempK))
            tempD[x] = tempK[x] * 1 / 3 + tempD[x] * 2 / 3
            D = np.row_stack((D, tempD))

    K = pd.DataFrame(K, index=Price.index, columns=Price.columns).round(2)
    D = pd.DataFrame(D, index=Price.index, columns=Price.columns).round(2)
    K_ = pd.DataFrame(K.unstack(), columns=['K'])
    D_ = pd.DataFrame(D.unstack(), columns=['D'])
    kdj = pd.concat([K_, D_], axis=1)
    kdj['J'] = kdj.K * 2 - kdj.D

def get_date_list(begin_date=None, end_date=None, year=1,pro=0):
    '''
    dates = get_date_list(datetime.date(2019, 2, 18), datetime.date.today())
    get_date_list()
    :param begin_date:
    :param end_date:
    :return:
    '''

    from pandas import read_csv
    import datetime
    import os
    import tushare as ts
    import dateutil.parser

    if not os.path.exists('e:\\yk\\cal_dates.csv'):
        cal_dates = ts.trade_cal()  # 返回交易所日历，类型为DataFrame, calendarDate  isOpen
        cal_dates.to_csv('e:\\yk\\cal_dates.csv', encoding='GBK')

    date_list = [];
    dates = []
    if end_date == None:
        end_date = datetime.date.today()
    elif isinstance(end_date, str):
        end_date = dateutil.parser.parse(end_date)
    if begin_date == None:
        begin_date = end_date - datetime.timedelta(days=365 * year)
    elif isinstance(begin_date, str):
        begin_date = dateutil.parser.parse(begin_date).date()
    # cal_dates = ts.trade_cal()

    cal_dates = read_csv('e:\\yk\\cal_dates.csv', encoding='GBK').iloc[:, 2:]
    dates = cal_dates[cal_dates.isOpen == 1]
    dates.drop('isOpen',axis=1,inplace=True)
    dates = dates[(str(begin_date) <= dates.calendarDate) & (dates.calendarDate <= str(end_date))]
    dates=dates.reset_index(drop=True)
    if pro==1:return dates #更多指标
    if pro==0: #普通下
        return dates.calendarDate.tolist()
    # while begin_date <= end_date:
    #     # date_str = str(begin_date)
    #     date_list.append(begin_date)
    #     begin_date += datetime.timedelta(days=1)
    # def is_open_day(date):
    #     if date in cal_dates['calendarDate'].values:
    #         return cal_dates[cal_dates['calendarDate'] == date].iat[0, 1] == 1
    #     return False
    # dates = [str(x) for x in date_list if is_open_day(str(x))]

def get_advanceDate(date=None, shift=-1):
    '''

    :param date:
    :param shift: 负数往过去几天，正数往以后几天
    :return:
    '''
    import datetime
    from pandas import read_csv
    if date == None: date = str(datetime.date.today())
    if not isinstance(date, str): date = str(date)
    # cal_dates = read_csv('e:\\yk\\cal_dates.csv', encoding='GBK').iloc[:, -2:]
    cal_dates = read_csv('e:\\yk\\cal_dates.csv', encoding='GBK').iloc[:, [2, 3]]
    cal_dates = cal_dates[cal_dates.isOpen == 1]['calendarDate']
    if shift < 0: return cal_dates[cal_dates < date].values[shift]
    if shift > 0: return cal_dates[cal_dates > date].values[shift - 1]
    if shift == 0: return cal_dates[cal_dates <= date].values[-1]

def get_stk_list(update=0):
    import tushare as ts
    from pandas import read_csv
    import os
    # 获取所有股票数据，利用股票代码获取复权数据
    if (not os.path.exists('e:\\yk\\stock_basics.csv')) | update:
        stock_basics = ts.get_stock_basics()
        stock_basics.sort_index(inplace=True)
        stock_basics.to_csv('e:\\yk\\stock_basics.csv', encoding='GBK')
    stock_basics = read_csv('e:\\yk\\stock_basics.csv', encoding='GBK')
    return stock_basics.code.astype('str').str.zfill(6).tolist()

def get_price(N_year=3, updata=0, adj=1, market_value=0):
    '''
    stk_list = DataAPI.EquGet(equTypeCD=u"A",field=["secID", u"ticker",'listDate','delistDate'],pandas="1")
    stk = stk_list[stk_list.delistDate!=stk_list.delistDate].ticker.tolist()
    for i in range(7):
    year = '201' +str(i)
    begin = year + '0101';
    end = year + '1231';
    df = DataAPI.MktEqudAdjGet(ticker=stk,tradeDate=u"",beginDate=begin,endDate=end,isOpen="",
                            field=u"ticker,tradeDate,closePrice",pandas="1")
    # df = df.set_index(['ticker','tradeDate']).unstack()
    # df.columns = df.columns.levels[1]
    df = df.pivot(index = 'tradeDate',columns='ticker',values='closePrice')
    df.to_csv(year + '.csv',encoding='gbk')

    :param N_year:自today的近N年,adj
           adj=1前复权 adj=0 不复权
    :return:
    '''
    import datetime
    import pandas as pd
    import numpy as np
    import urllib
    now = datetime.datetime.now().date()
    year = now.year
    now = get_advanceDate(now, 0)
    year_ = year + 1 - N_year
    df = pd.DataFrame()
    # i=2019
    for i in range(year_, year + 1):
        if market_value == 1:
            t = pd.read_csv('e:\\yk\\' + str(i) + '市值.csv').set_index('tradeDate')
        else:
            if adj:                  #前复权
                t = pd.read_csv('e:\\yk\\' + str(i) + '.csv').set_index('tradeDate')
            else:                     #除权
                t = pd.read_csv('e:\\yk\\' + str(i) + '_.csv').set_index('tradeDate')
        if i == year & updata == 1:
            enddate = t.index[-1]
            if enddate != now:
                dates = get_date_list(enddate, now)[1:]
                start_date = dates[0].replace('-', '')
                end_date = dates[-1].replace('-', '')
                stk = get_stk_list()
                stk = [str(1 - int(x[0] == '6')) + x for x in stk]
                dic = dict()

                def f(s):
                    global dic
                    url = 'http://quotes.money.163.com/service/chddata.html?code=' + s + '&start=' + start_date + '&end=' + end_date + '&fields=TCLOSE;HIGH;LOW;TOPEN;LCLOSE;CHG;PCHG;TURNOVER;VOTURNOVER;VATURNOVER;TCAP;MCAP'
                    try:
                        t = urllib.request.urlopen(url).read()
                        t = t.decode('gbk').split('\r\n')  # 该段获取原始数据
                        # col_info = t[0].split(',')  # 各列的含义:0日期1股票代码2名称3收盘价4最高价5最低价,开盘价,前收盘,涨跌额,涨跌幅,换手率,成交量,成交金额,总市值,流通市值
                        index_data = [t[iter].split(',')[3] for iter in range(1, len(t) - 1)]
                        index_data.reverse()
                        dic[s] = index_data
                    except:
                        print(s)

                thread(stk, None, f)

                for s in list(set(stk) - set(dic.keys())):
                    try:
                        f(s)
                    except:
                        print(s)
                [dic[x].extend((len(dates) - len(dic[x])) * [np.nan]) for x in dic.keys()]
                dic = pd.DataFrame(dic, index=dates, sort=True)
                dic.columns = dic.columns.str[1:]
                t = t.append(dic, sort=False)
                t.index.names = ['tradeDate']
                t.to_csv('e:\\yk\\' + str(i) + '.csv', encoding='gbk')

        df = df.append(t)
    df.dropna(how='all', axis=1, inplace=True)
    return df

def get_cons(indexname = 'ZZ500'):
    import pandas as pd
    t = pd.read_csv('e:\\yk\\'+indexname+'.csv').set_index('calendarDate')
    t=t.astype(str)
    t=t.apply(lambda x:x.str.zfill(6))
    return t

def get_factor(factor_name,before=0):
    '''
    ROE ROA PB PE
    :param factor:
    :return:
    '''
    import pandas as pd
    def _factor_handlle(factor):
        tradeDate = factor.iloc[:, 0][2:]
        factor = factor.iloc[:, factor.columns.str.contains(factor_name)]
        factor.columns = factor.iloc[0, :].str[:6]
        factor = factor[2:]
        factor.index = tradeDate
        factor = factor.astype('float')
        factor.index.name = 'tradeDate';
        factor.columns.name = 'ticker'
        return factor
    if factor_name in ['ROE','ROA','PE','PB']:
        factor = pd.read_csv('e:\\yk\\factor.csv')
        factor = _factor_handlle(factor)
        if before==1:
            factor1 = pd.read_csv('e:\\yk\\factor_before.csv')
            factor1 = _factor_handlle(factor1)
            factor = factor1.append(factor)
        return factor
    if factor_name in ['NetProfitGrowRate','NetAssetGrowRate','FEARNG','FSALESG']:
        factor = pd.read_csv('e:\\yk\\factor1.csv')

def get_index(index_name = '沪深300',length = 500):
    import pandas as pd
    factor = pd.read_csv('e:\\yk\\{}.csv'.format(index_name))
    factor = factor[['tradeDate','closeIndex']]
    return factor.set_index('tradeDate')['closeIndex'][-length:]

def get_industry():
    import pandas as pd
    t = pd.read_csv('e:\\yk\\industry.csv', encoding='gbk')
    t = t[['ticker', 'industryName1']]
    t = t.rename_axis({'industryName1': 'industry'}, axis=1)
    t = t[t.ticker.str.startswith('0') | t.ticker.str.startswith('3') | t.ticker.str.startswith('6')]
    t=t.reset_index(drop=True)
    return t

def formula(t):
    import sympy
    if isinstance(t, str):
        for i in range(10, 1, -1):
            try:
                t = eval(t)
                break
            except Exception as e:
                a = str(e).split("'")[1]
                exec(a + '=sympy.Symbol(a)')

    t = sympy.latex(t)
    t = t.center(len(t) + 2, '$')
    # list_i = list(t)    # str -> list
    # list_i.insert(0, '$')   # 注意不用重新赋值
    # list_i.insert(len(t)+1, '$')
    # str_i = ''.join(list_i)    # list -> str
    import matplotlib.pyplot as plt
    fig = plt.figure()
    ax = fig.add_subplot(111)
    ax.set_xlim([1, 10])
    ax.set_ylim([1, 5])
    ax.set_xticks([])
    ax.set_yticks([])
    ax.spines['left'].set_color('none')
    ax.spines['right'].set_color('none')
    ax.spines['bottom'].set_color('none')
    ax.spines['top'].set_color('none')
    ax.text(3, 3, t, fontsize=28)
    plt.show()

    # t='$\int_a^b f(x)\mathrm{d}x$'

def thread(stk, dates, func):
    from queue import Queue
    import threading
    import os
    import datetime
    import pandas as pd
    import tushare as ts

    stock_code_queue = Queue()
    if dates == None: dates = ['日期无需循环']
    for date in dates:  # 其实，这里不需要日期
        for code in stk:
            stock_code_queue.put((code, date))
    task_qeue = stock_code_queue

    # 获取复权数据
    def process_data(task_qeue):
        # queueLock.acquire()
        while not task_qeue.empty():
            t = task_qeue.get()
            s = t[0]
            print("正在获取%s;数据还有%s条:" % (s, task_qeue.qsize()))
            try:
                func(s)
            except:
                print(s + '  ' + date + '  None')

    class get_qfq(threading.Thread):
        def __init__(self, name, queue):
            threading.Thread.__init__(self)
            self.name = name
            self.queue = queue

        def run(self):
            process_data(self.queue)
            print("Exiting " + self.name)

    starttime = datetime.datetime.now()
    threads = []
    for i in range(20):
        thread = get_qfq('thread' + str(i), stock_code_queue)
        thread.start()
        threads.append(thread)
    for thread in threads:
        thread.join()
    endtime = datetime.datetime.now()
    print((endtime - starttime))

def add_sheet(data, sheetname,wookbookname=None):
    from openpyxl import load_workbook,Workbook
    import os
    from pandas import ExcelWriter
    if wookbookname.__contains__('.xlsx'):wookbookname=wookbookname[:-5]
    if not os.path.exists(wookbookname + '.xlsx'):
        wb = Workbook()
        wb.save(filename=wookbookname + '.xlsx')
    book = load_workbook(wookbookname + '.xlsx')
    writer = ExcelWriter(wookbookname + '.xlsx', engine='openpyxl')
    writer.book = book
    data.to_excel(writer, sheetname)
    writer.save()

def quantile(t,n_quantile = 10):          # 1是最小的，10是最大的
    import pandas as pd
    # 统计十分位数
    pct_quantiles = 1.0 / n_quantile

    rank = pd.Series(index=t.index)

    for i in range(n_quantile):
        if i==0:div_mean_results = []
        down = t.quantile(pct_quantiles * i)
        up = t.quantile(pct_quantiles * (i + 1))
        t1 = (t <= up) & (t >= down)
        rank[t1] = i+1
    rank.name=t.name
    return rank

